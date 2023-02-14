from django.shortcuts import render
from .models import Gallery,Team,logo,Carrer,blog,Testimonials,Events,Birac,Tbi,Investors,Govt_Tie,OurStartup,UploadImage,International_Partners,Sisfs,WhoAreWe,Contact_SECTION,HOME_TESTIMONIAL,EventsForm,MentorConnectDB,MentorClinicDB, angelInvestorDB, new_venturesDB
from .Tools import get_images,get_team,reguler_datas,get_blog
import datetime
import json
import openpyxl

# Create your views here.


def admin(request):
    item = get_images()
    return render(request,"pages/empty.html",reguler_datas({"categories":item[0],"images":item[1]}))


#...............gallery.......................................
def gallery(request):
    item = get_images()
    return render(request,"about_us/gallery.html",reguler_datas({"categories":item[0],"images":item[1]}))
#............................................................
# upload image...............................................
def upload_image(request):
    categories = request.POST.get("Category")
    image = request.FILES["image_file"]
    update = Gallery(image=image,categories=categories)
    update.save()
    return render(request,"about_us/team.html")

def delete_image(request):
    id = request.POST.get("id")
    image = Gallery.objects.get(G_id=id)
    image.delete()
    return render(request,"about_us/team.html")
#..............................................................

def aboutus(request):
    return render(request,"about_us/team.html",reguler_datas({"team":get_team()}))


# upload team ...............................
def admin_team(request):
    return render(request,'pages/team.html',reguler_datas({"team":get_team()}))

def update_team(request):
    name = request.POST.get("name")
    Category = request.POST.get("Category")
    position = request.POST.get("position")
    image = request.FILES["image_file"]
    db_obj = Team(Name=name,categories=Category,image=image,position=position)
    db_obj.save()
    obj = Team.objects.all()
    for i in obj:
        print(i.Name,i.image)

    return render(request,'pages/team.html')

def delete_team(request):
    id = request.POST.get("id")
    image = Team.objects.get(Team_id=id)
    image.delete()
    return render(request,"about_us/team.html",reguler_datas())

#............................................................
#...............Logo.........................................
def update_logo(request):
    return render(request,"home/logo.html",reguler_datas())

def upload_logo(request):
    Reson = request.POST.get("#reson")
    image = request.FILES["#fileInput-single"]
    update = logo(image=image,Reson=Reson)
    update.save()
    return render(request,"home/logo.html")

def delete_logo(request):
    id = request.POST.get("id")
    size = len(logo.objects.all())
    if size > 1:
        image = logo.objects.get(L_id=id)
        image.delete()
    else:
        pass
    return render(request,"home/logo.html")

def set_logo(request):
    id = request.POST.get("id")
    image = logo.objects.get(L_id=id)
    image_ = image.image
    reson = image.Reson
    image.delete()
    obj = logo(image=image_,Reson=reson)
    obj.save()
    return render(request,"home/logo.html")
#............................................................
#...............Logo.........................................

def carrer(request):
    return render(request,"about_us/carrer.html")
def update_carrer(request):
    ids=['#name','#Qualification','#Experience','#Email','#Subject','#Message']
    Name = request.POST.get(ids[0])
    Email = request.POST.get(ids[3])
    Message = request.POST.get(ids[-1])
    Subject = request.POST.get(ids[4])
    qualififcation = request.POST.get(ids[1])
    experience = request.POST.get(ids[2])
    image = request.FILES["#fileInput-single"]
    obj = Carrer(Name=Name,Email=Email,Message=Message,Subject=Subject,qualififcation=qualififcation,experience=experience,image=image)
    obj.save()
    ob=Carrer.objects.all()
    for i in ob:
        print(i.Name)
    return render(request,"about_us/carrer.html")
    
#............................................................

#...............Blog........................................
def blog_edit(request):
    return render(request,"pages/blog_edit.html")

def save_blog(request):
    ids = ['#title','#description','#content','#Category','#Thumbnail']
    title = request.POST.get(ids[0])
    description = request.POST.get(ids[1])
    content = request.POST.get(ids[2])
    Category = request.POST.get(ids[3])
    Thumbnail = request.POST.get(ids[4])

    obj = blog(title=title,description=description,content=content,categories=Category,blog_profile_img=Thumbnail)
    obj.save()
    ob = blog.objects.all()
    for i in ob:
        print(i.blog_profile_img,i.title,i.content)

    return render(request,"pages/blog_edit.html")

def save_edit_blog(request,pk):
    ids = ['#title','#description','#content','#Category','#Thumbnail']
    title = request.POST.get(ids[0])
    description = request.POST.get(ids[1])
    content = request.POST.get(ids[2])
    Category = request.POST.get(ids[3])
    Thumbnail = request.POST.get(ids[4])

    obj = blog.objects.get(id=pk)
    obj.content = content
    obj.title = title
    obj.description = description
    obj.categories = Category
    obj.blog_profile_img = Thumbnail
    obj.save()

    print("Saved...........")

    return render(request,"pages/blog_edit.html")


def list_blog(request):
    items = get_blog()
    return render(request,"home/blog.html",{'blogs':items})

def view_blog(request,pk):
    page = blog.objects.get(id=pk)
    items = get_blog()
    return render(request,"home/view_blog.html",{'blog':page,'item':items})

def delete_blog(request):
    bl_id = request.GET.get("id")
    page = blog.objects.get(id=bl_id)
    page.delete()
    return render(request,"home/view_blog.html",{'blog':page})


def list_edit_blog(request):
    items = get_blog()
    return render(request,"home/edit_blog_list.html",{'blogs':items})

def edit_blog(request,pk):
    obj = blog.objects.get(id=pk)
    return render(request,"pages/blog_re_edit.html",{'obj':obj})


#............................................................

#.....................Events.........................

def events(request):
    obj = Events.objects.all()
    return render(request,"about_us/Events.html",reguler_datas({"card":obj}))

def events_edit(request):
    obj = Events.objects.all()
    return render(request,"pages/Events_edit.html",reguler_datas({"card":obj}))


def events_save(request):
    ids = ['#topic','#program','#date','#Category','#fileInput-single']
    description = request.POST.get(ids[1])
    topic = request.POST.get(ids[0])
    date = request.POST.get(ids[2])
    categories = request.POST.get(ids[3])
    image = request.FILES[ids[-1]]
    d = date.split("-")
    date_formate = datetime.date(int(d[0]), int(d[1]), int(d[2]))
    obj = Events(description=description,type=topic,categories=categories,date=date_formate,image=image)
    obj.save()
    print("....................saved.......")
    return render(request,"about_us/Events.html",reguler_datas({"card":obj}))


#............................................................


#.....................Testimonicals.........................

def Testimonicals(request):
    obj = Testimonials.objects.all()
    return render(request,"about_us/Testimonicals.html",reguler_datas({"card":obj}))

def Testimonicals_edit(request):
    obj = Testimonials.objects.all()
    return render(request,"pages/Testimonicals_edit.html",reguler_datas({"card":obj}))

def Testimonicals_save(request):
    vals = ['#Name','#position','#description','#Category','#fileInput-single']
    Name = request.POST.get(vals[0])
    position = request.POST.get(vals[1])
    image = request.FILES[vals[-1]]
    description = request.POST.get(vals[2])
    categories = request.POST.get(vals[3])

    vals = Testimonials(Name=Name,position=position,image=image,description=description,categories=categories)
    vals.save()

    return render(request,"pages/Testimonicals_edit.html")

#............................................................

#...............birac........................................
def birac(request):
    obj = Birac.objects.latest('id')
    re_view = obj.overview.split(',')
    topic= {}
    for i in re_view:
        if '~' in i:
           topic[i.split('~')[1]] = i.split('~')[0]
            
    print(topic)

    return render(request,"about_us/birac.html",reguler_datas({'birac':obj,"topic":topic}))

def birac_edit(request):
    obj = Birac.objects.all()
    item = []
    topic= {}
    birac_ = {}
    for i in obj:
        topic=dict()
        for j in i.overview.split(','):
             if '~' in j:
                topic[j.split('~')[1]] = j.split('~')[0]
        item.append(topic)
    print(item)
    for i,x in enumerate(obj):
        birac_[x] = item[i]
    print(birac_)
    return render(request,"pages/birac_edit.html",reguler_datas({'birac':birac_,'view':item}))

def birac_save(request):
    ids = ['#title','#subtitle','#content','tags']
    title =request.POST.get(ids[0])
    subtitle_ = request.POST.get(ids[1])
    content = request.POST.get(ids[2])
    tags = request.POST.get(ids[3])
    item=""
    print(subtitle_)
    res = json.loads(tags)
    for i in res:
        item = item + i.get('value') + ", "
    obj = Birac(title=title,subtitle=subtitle_,description=content,overview=item)
    obj.save()
    print("/..........................................saved................................................../")
    return render(request,"pages/birac_edit.html",reguler_datas())

def delete_birac(request):
    bl_id = request.GET.get("id")
    page = Birac.objects.get(id=bl_id)
    page.delete()
    return render(request,"home/view_blog.html",{'blog':page})

def set_birac(request):
    id_ = request.GET.get("id")
    image = Birac.objects.get(id=id_)
    title = image.title
    subtitle = image.subtitle
    description = image.description
    overview = image.overview
    image.delete()
    obj = Birac(title=title,subtitle=subtitle,description=description,overview=overview)
    obj.save()
    print("saved")
    return render(request,"home/logo.html")

#............................................................

#...............tbi........................................
def tbi(request):
    try:
        obj = Tbi.objects.latest('id')
        re_view = obj.overview.split(',')
        topic= {}
        for i in re_view:
            if '~' in i:
                topic[i.split('~')[1]] = i.split('~')[0]
    except:
        obj = ""
        topic = ""
        print("maybe the database are empty.....")
            
    print(topic)

    return render(request,"about_us/tbi.html",reguler_datas({'birac':obj,"topic":topic}))

def tbi_edit(request):
    obj = Tbi.objects.all()
    item = []
    topic= {}
    tbi_ = {}
    for i in obj:
        topic=dict()
        for j in i.overview.split(','):
             if '~' in j:
                topic[j.split('~')[1]] = j.split('~')[0]
        item.append(topic)
    print(item)
    for i,x in enumerate(obj):
        tbi_[x] = item[i]
    print(tbi_)
    return render(request,"pages/tbi_edit.html",reguler_datas({'birac':tbi_,'view':item}))

def tbi_save(request):
    ids = ['#title','#subtitle','#content','tags']
    title =request.POST.get(ids[0])
    subtitle_ = request.POST.get(ids[1])
    content = request.POST.get(ids[2])
    tags = request.POST.get(ids[3])
    item=""
    print(subtitle_)
    res = json.loads(tags)
    for i in res:
        item = item + i.get('value') + ", "
    obj = Tbi(title=title,subtitle=subtitle_,description=content,overview=item)
    obj.save()
    print("/..........................................saved................................................../")
    return render(request,"pages/tbi_edit.html",reguler_datas())

def delete_tbi(request):
    bl_id = request.GET.get("id")
    page = Tbi.objects.get(id=bl_id)
    page.delete()
    return render(request,"home/view_blog.html",{'blog':page})

def set_tbi(request):
    id_ = request.GET.get("id")
    image = Tbi.objects.get(id=id_)
    title = image.title
    subtitle = image.subtitle
    description = image.description
    overview = image.overview
    image.delete()
    obj = Tbi(title=title,subtitle=subtitle,description=description,overview=overview)
    obj.save()
    print("saved")
    return render(request,"home/logo.html")

#............................................................

#...............sisfs........................................
def sisfs(request):
    try:
        obj = Sisfs.objects.latest('id')
        re_view = obj.overview.split(',')
        topic= {}
        for i in re_view:
            if '~' in i:
                topic[i.split('~')[1]] = i.split('~')[0]
    except:
        obj = ""
        topic = ""
        print("maybe the database are empty.....")
            
    print(topic)

    return render(request,"about_us/sisfs.html",reguler_datas({'birac':obj,"topic":topic}))

def sisfs_edit(request):
    obj = Sisfs.objects.all()
    item = []
    topic= {}
    sisfs_ = {}
    for i in obj:
        topic=dict()
        for j in i.overview.split(','):
             if '~' in j:
                topic[j.split('~')[1]] = j.split('~')[0]
        item.append(topic)
    print(item)
    for i,x in enumerate(obj):
        sisfs_[x] = item[i]
    print(sisfs_)
    return render(request,"pages/sisfs_edit.html",reguler_datas({'birac':sisfs_,'view':item}))

def sisfs_save(request):
    ids = ['#title','#subtitle','#content','tags']
    title =request.POST.get(ids[0])
    subtitle_ = request.POST.get(ids[1])
    content = request.POST.get(ids[2])
    tags = request.POST.get(ids[3])
    item=""
    print(subtitle_)
    res = json.loads(tags)
    for i in res:
        item = item + i.get('value') + ", "
    obj = Sisfs(title=title,subtitle=subtitle_,description=content,overview=item)
    obj.save()
    print("/..........................................saved................................................../")
    return render(request,"pages/sisfs_edit.html",reguler_datas())

def delete_sisfs(request):
    bl_id = request.GET.get("id")
    page = Sisfs.objects.get(id=bl_id)
    page.delete()
    return render(request,"home/view_blog.html",{'blog':page})

def set_sisfs(request):
    id_ = request.GET.get("id")
    image = Sisfs.objects.get(id=id_)
    title = image.title
    subtitle = image.subtitle
    description = image.description
    overview = image.overview
    image.delete()
    obj = Sisfs(title=title,subtitle=subtitle,description=description,overview=overview)
    obj.save()
    print("saved")
    return render(request,"home/logo.html")

def eventform(request):
    return render(request,"about_us/event_form.html")

def edit_eventform(request):
    obj = EventsForm.objects.all()
    return render(request,"pages/eventform_edit.html",{'obj':obj})

def update_eventform(request):
    title = request.POST.get('#title')
    Name = request.POST.get('#name')
    Email = request.POST.get('#Email')
    company = request.POST.get('#company')
    event = request.POST.get('#event')
    linkedin = request.POST.get('#linkedin')
    website  =request.POST.get('#website')
    image = request.FILES["#fileInput-single"]
    obj = EventsForm(title=title,Name=Name,Email=Email,linkedin=linkedin,website=website,company=company,event=event,image=image)
    obj.save()
    ob=EventsForm.objects.all()
    for i in ob:
        print(i.Name)
    return render(request,"about_us/carrer.html")

def delete_form(request):
    bl_id = request.GET.get("id")
    page = EventsForm.objects.get(id=bl_id)
    page.delete()
    return render(request,"home/view_blog.html",{'blog':page})

def convert_excel(request):
    wb = openpyxl.Workbook() 
    sheet = wb.active 
    data = EventsForm.objects.all()
    title = ["updated_date","title","Name","Email","company","event","linkedin","website"]
    for i,x in enumerate(title):
        cell_obj = sheet.cell(row = 1, column = i+1)
        cell_obj.value = x

    for i,x in enumerate(data,2):
        row_data = [x.updated_date,x.title,x.Name,x.Email,x.company,x.event,x.linkedin,x.website]
        for j,y in enumerate(row_data):
            cell_obj = sheet.cell(row = i+1, column = j+1)
            cell_obj.value = y
            print(i,j)
    wb.save("sample5.xlsx") 
    return render(request,"home/view_blog.html")


def EDI (request):
    return render(request,"fund/edi.html")



def about (request):
    return render(request,"a/thiruabout.html")

def angelInvestor (request):
    return render(request,"angelinvestor.html")

def GlobalMarket (request):
    return render(request,"globalmarket.html")

def home(request):
    return render(request,"index.html")

def MentorClinic (request):
    return render(request,"mentorclinic.html")



def MentorConnect (request):
    return render(request,"mentorconnect.html",{'mentor':MentorConnectDB.objects.all()[::-1]})

def MontorConnect_edit(request):
    return render(request,"edtior/mentor_connect_edit.html",{'mentor':MentorConnectDB.objects.all()[::-1]})

def MontorConnect_save(request):
    content = request.POST.get('#content')
    if request.POST.get('#page') == 'MentorConnect':
        obj = MentorConnectDB(Content=content)
        obj.save()
    print("saved.....................................////////////////////////")
    return render(request,"edtior/mentor_connect_edit.html")



def MentorClinic (request):
    return render(request,"mentorclinic.html")

def Mentor_Clinic_edit(request):
    return render(request,"edtior/Mentor_Clinic_edit.html",{'mentor':MentorClinicDB.objects.all()[::-1]})

def Mentor_Clinic_save(request):
    content = request.POST.get('#content')
    obj = MentorClinicDB(Content=content)
    obj.save()
    print("saved.....................................////////////////////////")
    return render(request,"edtior/Mentor_Clinic_edit.html")


def angelInvestor (request):
    return render(request,"angelInvestor.html",{'mentor':angelInvestorDB.objects.all()[::-1]})

def angelInvestor_edit(request):
    return render(request,"edtior/angelInvestor_edit.html",{'mentor':angelInvestorDB.objects.all()[::-1]})

def angelInvestor_save(request):
    content = request.POST.get('#content')
    obj = angelInvestorDB(Content=content)
    obj.save()
    for i in angelInvestorDB.objects.all():
        print(i.Content)
    print("saved.....................................////////////////////////")
    return render(request,"edtior/angelInvestor_edit.html")

def new_ventures (request):
    data = new_venturesDB.objects.all()[::-1]
    print(data)
    return render(request,"newventures.html",{'mentor':data,'sample':'hi'})

def new_ventures_edit(request):
    return render(request,"edtior/new_ventures_edit.html",{'mentor':new_venturesDB.objects.all()[::-1]})

def new_ventures_save(request):
    content = request.POST.get('#content')
    obj = new_venturesDB(Content=content)
    obj.save()
    for i in new_venturesDB.objects.all():
        print(i.Content)
    print("saved.....................................////////////////////////")
    return render(request,"edtior/new_ventures_edit.html")

def ourStartups (request):
    return render(request,"ourstartup.html")

def samridth (request):
    return render(request,"samridth.html")

def service (request):
    return render(request,"service.html")

def sisfs(request):
    return render(request,"sisfs.html")

def stategovtfunds (request):
    return render(request,"stategovtfund.html")

def testimonial (request):
    return render(request,"testimonials.html")

def Mba (request):
    return render(request,"mba.html")

def career (request):
    return render(request,"carrer.html")

def gallery (request):
    return render(request,"gallery.html")


def home(request):
    try :
        whoweare = WhoAreWe.objects.all()[::-1]
        home_TESTIMONIAL =  HOME_TESTIMONIAL.objects.all()[::-1]
        contact_Section =  Contact_SECTION.objects.all()[::-1]
        investors = Investors.objects.all()[::-1]
        govt = Govt_Tie.objects.all()[::-1]
        Uploadimage = UploadImage.objects.all()[::-1]

        
        Internationalpartners = International_Partners.objects.all()[::-1]
        return render(request,"index.html",{'whoweare':whoweare,'ht':home_TESTIMONIAL,'cs':contact_Section,'investors':investors,'ip':Internationalpartners,'govt':govt,'Uploadimage':Uploadimage})
    except:
        print("maybe database are empty")
    return render(request,"pages/home_edit.html")


def home_edit(request):
    try :
        whoweare = WhoAreWe.objects.all()[::-1]
        home_TESTIMONIAL =  HOME_TESTIMONIAL.objects.all()[::-1]
        contact_Section =  Contact_SECTION.objects.all()[::-1]
        investors = Investors.objects.all()[::-1]
        Internationalpartners = International_Partners.objects.all()[::-1]
        govt = Govt_Tie.objects.all()[::-1]
        Uploadimage = UploadImage.objects.all()[::-1]


        return render(request,"pages/home_edit.html",{'whoweare':whoweare,'ht':home_TESTIMONIAL,'cs':contact_Section,'investors':investors,'ip':Internationalpartners,'govt':govt,'Uploadimage':Uploadimage})
    except:
        print("maybe database are empty")
    return render(request,"pages/home_edit.html")


def Whoweare(request):
    ids = ['#Sub-heading','#Point-1','#Point-2','#Point-3','#Point-4','#image_file']
    SubHeading = request.POST.get(ids[0])
    Point1 = request.POST.get(ids[1])
    Point2 = request.POST.get(ids[2])
    Point3 = request.POST.get(ids[3])
    Point4 = request.POST.get(ids[4])

    try :
        image = request.FILES[ids[5]]
    except:
        image = WhoAreWe.objects.all()[::-1][0].image

    obj = WhoAreWe(SubHeading=SubHeading,Point1=Point1,Point2=Point2,Point3=Point3,Point4=Point4,image=image)
    obj.save()

    for i in WhoAreWe.objects.all():
        print(i.SubHeading)

    return render(request,"gallery.html")


def  Home_TESTIMONIAL(request):
    ids = ['#fileInput-single1','#Testimonial','#Name','#Designation']

    Testimonial_content = request.POST.get(ids[1])
    Name = request.POST.get(ids[2])
    Designation = request.POST.get(ids[3])
    try :
        image = request.FILES[ids[0]]
    except:
        image = HOME_TESTIMONIAL.objects.all()[::-1][0].image

    obj = HOME_TESTIMONIAL(Testimonial_content=Testimonial_content,Name=Name,Designation=Designation,image=image)
    obj.save()

    for i in HOME_TESTIMONIAL.objects.all():
        print(i.Name)

    return render(request,"gallery.html")

def  Contact_Section(request):
    ids = ['#Title','#Address','#pno','#E-Mail']

    Title = request.POST.get(ids[0])
    Address = request.POST.get(ids[1])
    Phone_number = request.POST.get(ids[2])
    E_Mail = request.POST.get(ids[3])

    obj = Contact_SECTION(Title=Title,Address=Address,Phone_number=Phone_number,E_Mail=E_Mail)
    obj.save()

    for i in Contact_SECTION.objects.all():
        print(i.Title)

    return render(request,"gallery.html")


def investors(request):
    ids = ['#fileInput-single2','#sub-heading']

    heading = request.POST.get(ids[1])
    try :
        image = request.FILES[ids[0]]
        obj = Investors(Title=heading,image=image)
        obj.save()
    except:
        image = Investors.objects.all()[::-1][0]
        image.Title = heading
        image.save()


    for i in Investors.objects.all():
        print(i.Title)

    return render(request,"gallery.html")

def International(request):
    ids = ['#fileInput-single3','#sub-heading6']

    heading = request.POST.get(ids[1])
    try :
        image = request.FILES[ids[0]]
        obj = International_Partners(Title=heading,image=image)
        obj.save()
    except:
        image = International_Partners.objects.all()[::-1][0]
        image.Title = heading
        image.save()


    for i in International_Partners.objects.all():
        print(i.Title)

    return render(request,"gallery.html")

def GovtTie(request):
    ids = ['#fileInput-single4','61']

    heading = request.POST.get(ids[1])
    try :
        image = request.FILES[ids[0]]
        obj = Govt_Tie(Title=heading,image=image)
        obj.save()
    except:
        image = Govt_Tie.objects.all()[::-1][0]
        image.Title = heading
        image.save()

    for i in Govt_Tie.objects.all():
        print(i.Title)

    return render(request,"gallery.html")

def Upload_Image(request):
    ids = ['#fileInput-single8','#Appear','#Appear-image']

    heading = request.POST.get(ids[1])
    content = request.POST.get(ids[2])

    try :
        image = request.FILES[ids[0]]
        obj = UploadImage(Title=heading,image=image,content=content)
        obj.save()
    except:
        image = UploadImage.objects.all()[::-1][0]
        image.Title = heading
        image.save()

    for i in UploadImage.objects.all():
        print(i.Title)

    return render(request,"gallery.html")

def save_ourstartups (request):
    return render(request,"ourstartup.html")

